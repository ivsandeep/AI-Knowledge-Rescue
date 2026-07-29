from openpyxl import load_workbook


class ExcelProcessor:

    def extractText(self, filePath):

        workbook = load_workbook(filePath)

        extractedText = ""

        for sheet in workbook.worksheets:

            extractedText += f"Sheet : {sheet.title}\n"

            for row in sheet.iter_rows():

                rowText = []

                for cell in row:

                    if cell.value is not None:
                        rowText.append(str(cell.value))

                extractedText += " | ".join(rowText) + "\n"

            extractedText += "\n"

        workbook.close()

        return extractedText